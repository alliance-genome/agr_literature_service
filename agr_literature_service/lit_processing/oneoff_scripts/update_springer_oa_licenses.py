"""
update_springer_oa_licenses.py
==============================
Update Resource table to add CC BY license for Springer Nature fully open access
journals that are currently missing license information.

This script fetches the list of Springer Nature fully open access journals from:
https://cms-resources.apps.public.k8s.springernature.io/springer-cms/rest/v1/content/27820860/data/v2

It then queries the database to find resources matching those ISSNs and updates
resources that are missing license information to set:
- license_list = ['CC BY']
- copyright_license_id = 1 (CC BY)

Usage:
    python update_springer_oa_licenses.py [--dry-run]

Options:
    --dry-run    Show what would be updated without making changes
"""
import argparse
import io
import logging
import sys
from os import path
from typing import Dict, List, Optional, Tuple

import requests
from sqlalchemy.orm import Session

from agr_literature_service.lit_processing.utils.sqlalchemy_utils import create_postgres_session
from agr_literature_service.api.models import ResourceModel, CrossReferenceModel
from agr_literature_service.api.user import set_global_user_id

logging.basicConfig(format="%(message)s")
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# Springer Nature fully open access journal list URL
SPRINGER_OA_URL = "https://cms-resources.apps.public.k8s.springernature.io/springer-cms/rest/v1/content/27820860/data/v2"

# CC BY license ID (most common Springer Nature OA license)
CC_BY_LICENSE_ID = 1
CC_BY_LICENSE_NAME = "CC BY"


def _find_column_index(headers: Dict[str, int], column_names: List[str]) -> Optional[int]:
    """Find column index by trying multiple possible column names."""
    for col_name in column_names:
        if col_name in headers:
            return headers[col_name]
    return None


def _parse_excel_headers(sheet) -> Tuple[Optional[Dict[str, int]], Optional[int]]:
    """Parse Excel sheet to find header row and column indices."""
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not row or not any(row):
            continue
        # Check if this looks like a header row
        row_values = [str(cell).strip().lower() if cell else '' for cell in row]
        if any('issn' in val or 'title' in val for val in row_values):
            headers: Dict[str, int] = {}
            for col_idx, cell in enumerate(row):
                if cell:
                    headers[str(cell).strip().lower()] = col_idx
            return headers, row_idx
    return None, None


def _parse_excel_data(sheet, header_row: int, eissn_col: int, title_col: int) -> Dict[str, str]:
    """Parse data rows from Excel sheet."""
    journals: Dict[str, str] = {}
    max_col = max(eissn_col, title_col)

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_idx <= header_row:
            continue
        if not row or len(row) <= max_col:
            continue

        eissn = row[eissn_col]
        title = row[title_col]

        if eissn and title:
            eissn_str = str(eissn).strip()
            title_str = str(title).strip()
            if eissn_str and title_str:
                # Normalize ISSN (remove hyphens)
                normalized_issn = eissn_str.replace('-', '')
                journals[normalized_issn] = title_str

    return journals


def fetch_springer_oa_journals() -> Dict[str, str]:
    """
    Fetch the list of Springer Nature fully open access journals from the API.

    The API returns an Excel file (.xlsx) containing journal information.

    Returns:
        Dictionary mapping normalized ISSN (no hyphens) to journal title
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl is required to read Excel files. Install with: pip install openpyxl")
        return {}

    logger.info(f"Fetching Springer Nature OA journal list from: {SPRINGER_OA_URL}")

    try:
        response = requests.get(SPRINGER_OA_URL, timeout=60)
        response.raise_for_status()
    except requests.RequestException as e:
        logger.error(f"Failed to fetch Springer OA journal list: {e}")
        return {}

    try:
        workbook = load_workbook(filename=io.BytesIO(response.content), read_only=True)
        sheet = workbook.active

        headers, header_row = _parse_excel_headers(sheet)
        if not headers or header_row is None:
            logger.error("Could not find header row in Excel file")
            return {}

        # Find eISSN and title column indices
        eissn_col = _find_column_index(headers, ['eissn', 'e-issn', 'electronic issn', 'issn (electronic)'])
        title_col = _find_column_index(headers, ['journal title', 'title', 'journal', 'journal name'])

        if eissn_col is None or title_col is None:
            logger.error(f"Could not find required columns. Found headers: {list(headers.keys())}")
            return {}

        journals = _parse_excel_data(sheet, header_row, eissn_col, title_col)
        workbook.close()

    except Exception as e:
        logger.error(f"Failed to parse Excel content: {e}")
        return {}

    logger.info(f"Found {len(journals)} Springer Nature OA journals")
    return journals


def find_resources_by_issn(db: Session, issn_to_title: Dict[str, str]) -> List[Tuple[int, str, str]]:
    """
    Query the database to find resources matching the given ISSNs.

    Args:
        db: Database session
        issn_to_title: Dictionary mapping normalized ISSN to journal title

    Returns:
        List of tuples (resource_id, title, eissn) for resources missing OA license
    """
    resources: List[Tuple[int, str, str]] = []

    # Query resources with ISSN cross-references
    query = db.query(
        ResourceModel.resource_id,
        ResourceModel.title,
        ResourceModel.license_list,
        CrossReferenceModel.curie
    ).join(
        CrossReferenceModel,
        ResourceModel.resource_id == CrossReferenceModel.resource_id
    ).filter(
        CrossReferenceModel.curie_prefix == 'ISSN',
        CrossReferenceModel.is_obsolete.is_(False)
    )

    for resource_id, title, license_list, curie in query:
        # Extract and normalize ISSN from curie (format: ISSN:1234-5678)
        issn = curie.replace('ISSN:', '').replace('-', '')

        if issn in issn_to_title:
            # Check if already has OA license
            if license_list and len(license_list) > 0:
                has_oa = any(
                    'cc' in lic.lower() or 'creative commons' in lic.lower() or 'open access' in lic.lower()
                    for lic in license_list if lic
                )
                if has_oa:
                    logger.debug(f"Skipping resource {resource_id} ({title[:40]}...) - already has OA license")
                    continue

            # Format ISSN with hyphen for display
            formatted_issn = f"{issn[:4]}-{issn[4:]}" if len(issn) == 8 else issn
            resources.append((resource_id, title or issn_to_title[issn], formatted_issn))

    # Remove duplicates (same resource may have multiple ISSNs)
    seen_ids = set()
    unique_resources = []
    for res in resources:
        if res[0] not in seen_ids:
            seen_ids.add(res[0])
            unique_resources.append(res)

    return unique_resources


def update_resources(db: Session, resources: List[Tuple[int, str, str]], dry_run: bool = False) -> None:
    """
    Update resources to add CC BY license information.

    Args:
        db: Database session
        resources: List of (resource_id, title, eissn) tuples
        dry_run: If True, don't commit changes
    """
    updated_count = 0
    error_count = 0

    for resource_id, title, eissn in resources:
        try:
            resource = db.query(ResourceModel).filter(
                ResourceModel.resource_id == resource_id
            ).first()

            if not resource:
                logger.warning(f"Resource not found: ID={resource_id}, Title={title}")
                error_count += 1
                continue

            # Update resource with CC BY license
            if not dry_run:
                resource.license_list = [CC_BY_LICENSE_NAME]
                resource.copyright_license_id = CC_BY_LICENSE_ID
                db.add(resource)

            updated_count += 1
            action = "Would update" if dry_run else "Updated"
            logger.info(f"{action}: ID={resource_id}, {title[:60]} (eISSN: {eissn})")

        except Exception as e:
            error_count += 1
            logger.error(f"Error updating resource ID={resource_id}: {e}")

    # Commit all changes
    if not dry_run:
        db.commit()
        logger.info("Changes committed to database.")
    else:
        logger.info("DRY RUN - No changes were made.")

    logger.info("=" * 60)
    logger.info(f"Summary: Updated={updated_count}, Errors={error_count}")


def main():
    parser = argparse.ArgumentParser(
        description="Update Springer Nature OA journals with CC BY license"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Show what would be updated without making changes"
    )

    args = parser.parse_args()

    logger.info("Springer Nature OA License Update Script")
    logger.info(f"Source: {SPRINGER_OA_URL}")
    logger.info(f"Dry run: {args.dry_run}")
    logger.info("=" * 60)

    # Fetch Springer Nature OA journal list
    issn_to_title = fetch_springer_oa_journals()
    if not issn_to_title:
        logger.error("Failed to fetch Springer OA journal list. Exiting.")
        sys.exit(1)

    # Connect to database
    db = create_postgres_session(False)

    # Set script user for audit trail
    script_name = path.basename(__file__).replace(".py", "")
    set_global_user_id(db, script_name)

    try:
        # Find resources matching the ISSNs that need updating
        resources = find_resources_by_issn(db, issn_to_title)
        if not resources:
            logger.info("No resources need updating. All matching journals already have OA licenses.")
            return

        logger.info(f"Found {len(resources)} resources to update")
        update_resources(db, resources, dry_run=args.dry_run)
    finally:
        db.close()


if __name__ == "__main__":
    main()
